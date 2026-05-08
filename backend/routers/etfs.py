"""ETF 마스터 + PDF(구성종목) API.

데이터 소스: `data/etf_info.xlsx` — 매일 사용자가 받아서 떨굼.
시트 1 (ETF 기본정보): 코드, 종목명, CU단위
시트 2 (ETF PDF내역): 종목코드, 주식수 (특수종목 A000000은 현금)

캐싱: 첫 요청 시 로드, mtime 비교로 자동 reload (배당/휴일과 동일 패턴).
종목코드는 'A' 접두 제거하여 6자리로 정규화 (futures_master와 통일).
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook

router = APIRouter(prefix="/etfs", tags=["etfs"])

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
ETF_FILE = DATA_DIR / "etf_info.xlsx"

CASH_CODE = "000000"  # A000000 = 원화현금 (PDF 시트 특수 행)

# 자체 NAV/fNAV 산출 불가 ETF 키워드 (종목명 기준).
# 레버리지/인버스: PDF에 지수선물·스왑이 들어가 일중 평가 불가 (전일정산가 필요 또는 PDF 자체가 빈 껍데기).
# 채권/혼합/통화/원자재/부동산 그룹은 별도로 그룹 prefix로 차단.
_NON_ARBITRABLE_NAME_KEYWORDS = ("레버리지", "인버스", "2X", "2x", "3X", "3x", "선물")


def _is_arbitrable(group: Optional[str], name: Optional[str]) -> bool:
    """ETF가 'PDF 현물 바스켓 vs 개별주식선물' 차익 거래 대상인지.
    False면 fNAV/실집행/차익BP 컬럼 의미 없음 → 프론트에서 흐림 처리.
    """
    g = (group or "").strip()
    n = (name or "")
    # 그룹이 비어있거나 주식 계열이 아니면 차익 불가 (채권/혼합/통화/원자재/부동산/기타).
    if not g.startswith("주식-"):
        return False
    # 주식 그룹 안에 섞여 있는 레버리지/인버스/2X.
    if any(k in n for k in _NON_ARBITRABLE_NAME_KEYWORDS):
        return False
    return True


class _Cache:
    src_mtime: float = 0.0
    etfs: dict[str, dict] = {}        # code → {name, cu_unit, group, lp, arbitrable}
    pdfs: dict[str, dict] = {}        # code → {as_of, stocks: [...], cash}
    loaded_at: Optional[str] = None


_cache = _Cache()


def _norm_code(code) -> str:
    """'A0000J0' → '0000J0', '000370' → '000370'. None은 빈 문자열."""
    if code is None:
        return ""
    s = str(code).strip().upper()
    if len(s) == 7 and s.startswith("A"):
        return s[1:]
    return s


def _ensure_loaded() -> None:
    if not ETF_FILE.exists():
        raise HTTPException(status_code=503, detail=f"ETF 데이터 파일 없음: {ETF_FILE.name}")
    src_mtime = ETF_FILE.stat().st_mtime
    if _cache.src_mtime == src_mtime and _cache.etfs:
        return

    try:
        wb = load_workbook(ETF_FILE, data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ETF 엑셀 파싱 실패: {e}")

    etfs = _parse_master(wb["ETF 기본정보"])
    pdfs = _parse_pdfs(wb["ETF PDF내역"])
    wb.close()

    _cache.etfs = etfs
    _cache.pdfs = pdfs
    _cache.src_mtime = src_mtime
    _cache.loaded_at = datetime.fromtimestamp(src_mtime).isoformat(timespec="seconds")


def _parse_master(ws) -> dict[str, dict]:
    """ETF 기본정보 시트 → {code: {name, cu_unit, group, lp}}.
    NAV/현재가/거래량은 조회 시점값이라 무시 — 실시간 데이터로 대체.
    """
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}
    idx = {h: i for i, h in enumerate(header) if h is not None}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out: dict[str, dict] = {}
    for row in rows:
        code = _norm_code(col(row, "코드"))
        if not code:
            continue
        name = str(col(row, "종목명") or "").strip()
        group = str(col(row, "그룹") or "").strip() or None
        out[code] = {
            "code": code,
            "name": name,
            "cu_unit": col(row, "CU단위"),
            "group": group,
            "lp": str(col(row, "LP1") or "").strip() or None,
            "arbitrable": _is_arbitrable(group, name),
        }
    return out


def _parse_pdfs(ws) -> dict[str, dict]:
    """ETF PDF내역 시트 → {etf_code: {as_of, stocks: [...], cash}}.
    A000000(원화현금)은 별도 cash 필드로 분리 (전일평가금액 사용).
    """
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}
    idx = {h: i for i, h in enumerate(header) if h is not None}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out: dict[str, dict] = {}
    for row in rows:
        etf_code = _norm_code(col(row, "ETF코드"))
        if not etf_code:
            continue
        stock_code = _norm_code(col(row, "종목코드"))
        if not stock_code:
            continue

        as_of = col(row, "날짜")
        as_of_str = as_of.date().isoformat() if isinstance(as_of, datetime) else (
            as_of.isoformat() if isinstance(as_of, date) else None
        )

        bucket = out.setdefault(etf_code, {"as_of": as_of_str, "stocks": [], "cash": 0})

        if stock_code == CASH_CODE:
            cash = col(row, "전일평가금액") or 0
            try:
                bucket["cash"] = int(cash)
            except (TypeError, ValueError):
                bucket["cash"] = 0
            continue

        qty = col(row, "주식수") or 0
        try:
            qty_int = int(qty)
        except (TypeError, ValueError):
            qty_int = 0

        bucket["stocks"].append({
            "code": stock_code,
            "name": str(col(row, "종목명") or "").strip(),
            "qty": qty_int,
        })
    return out


@router.get("")
async def list_etfs():
    """ETF 마스터 목록."""
    _ensure_loaded()
    items = sorted(_cache.etfs.values(), key=lambda d: d["code"])
    return {
        "loaded_at": _cache.loaded_at,
        "count": len(items),
        "items": items,
    }


@router.get("/pdf-all")
async def get_all_pdfs():
    """모든 ETF의 PDF를 한 번에. 스크리너 페이지가 1회 fetch로 다 받음."""
    _ensure_loaded()
    out = {}
    for code, pdf in _cache.pdfs.items():
        meta = _cache.etfs.get(code, {})
        out[code] = {
            "code": code,
            "name": meta.get("name"),
            "cu_unit": meta.get("cu_unit"),
            "arbitrable": meta.get("arbitrable", True),
            "as_of": pdf["as_of"],
            "cash": pdf["cash"],
            "stocks": pdf["stocks"],
        }
    return {
        "loaded_at": _cache.loaded_at,
        "count": len(out),
        "items": out,
    }


@router.get("/{code}/pdf")
async def get_etf_pdf(code: str):
    """ETF PDF 구성종목."""
    _ensure_loaded()
    norm = _norm_code(code)
    pdf = _cache.pdfs.get(norm)
    if not pdf:
        raise HTTPException(status_code=404, detail=f"PDF 없음: {code}")
    meta = _cache.etfs.get(norm, {})
    return {
        "code": norm,
        "name": meta.get("name"),
        "cu_unit": meta.get("cu_unit"),
        "arbitrable": meta.get("arbitrable", True),
        "as_of": pdf["as_of"],
        "cash": pdf["cash"],
        "stocks": pdf["stocks"],
    }
