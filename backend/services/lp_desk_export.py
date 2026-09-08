"""LP 데스크 파라미터 엑셀 내보내기 — lp-system-design.md §14.11 (내부망 반입용).

**왜 있나.** 실집행은 LENS가 없는 **내부망**에서 한다. 내부망에는 지수선물 wire조차 없어
(§14.11 말미 선결 과제) 버킷 계약수 환산이 화면에서 안 나온다. 그래서 외부망 LENS가 산출한
**정적 파라미터(β·호가 밴드·전일종가·CU)만 엑셀로 넘기고**, 매 초 변하는 값(체결 수량·현재가·
선물가)은 내부망 엑셀이 **DDE로 직접 받아** 워크북 안의 수식으로 헤지 수량까지 계산하게 한다.

    LENS(외부망)  →  파라미터 스냅샷 xlsx  →  내부망 엑셀 + DDE 시세  →  계약수 액션

**매크로 없음 — 순수 수식만.** 내부망 보안 정책상 VBA가 붙은 파일은 반입 자체가 막힌다.
따라서 계산은 전부 워크시트 수식(VLOOKUP/SUM/ROUND/IF)으로 표현하고, 파이썬은 값과 수식
문자열만 써 넣는다.

시트 5장:
  베타    36행 스냅샷 (/master 캐시 그대로). **코드는 텍스트 서식** — `0117V0`·`0052D0`처럼
          영숫자 코드가 있고, 숫자로 저장되면 VLOOKUP 키가 어긋난다.
          잔차σ 옆에 **오버나이트 상한·금지**(§14.12) — 내부망에서 14:30 정리 판단의 근거.
  OMS     OMS 조건변수 **A~D**(함수 전략 v1.5 · `docs/lp-oms-strategy-v1.html`)를 종목별로
          **그대로 옮겨 적는 표** (§14.11). 값은 베타 시트와 같은 μ·σ·z 소스에서 나온다 —
          파일 안에 숫자가 두 벌 생기지 않게 한 함수를 공유하고, **단위만 bp→%(÷100)** 로 바꾼다.
  체결    입력/DDE 200행. 코드→종목명·β를 VLOOKUP으로 끌어와 K200/KQ 노출을 행마다 계산.
  헤지    노출 합계 → 목표 계약수 → **집행할 계약(목표−보유)**. 선물가 미입력 시 DIV/0 가드.
  PARAMS  생성 시각·통계일·z·지평·승수·장중 상한 + 사용법. 승수는 헤지 시트가 여기를 참조한다.

호가 밴드 x는 프론트 `lib/lp-desk.ts::suggestQuote`와 **같은 산식**으로 서버에서 미리 채운다
(내부망에는 캘리브가 없으므로 값으로 굳혀 보낸다):

    x매도 = μ_g + z·√(σ_g² + σ_r²)   ·   x매수 = μ_g − z·σ_comb   ·   σ_r = s_diff_sigma_bp[T]

OMS 시트의 A·B·D는 **이 bp 값을 ÷100 한 %** 다 (OMS 조건변수는 % 부호 입력 — 매도 +5bp → 0.05).
정리 레벨·임계는 없다(v1.5) — 호가는 항상 진입 밴드 − 연속 편향(D×포지션/C) 하나고, 엑셀은
그 강도인 E(σ 배수)만 넘긴다. 산식이 두 벌 생기지 않게 bp 계산은 위 한 곳에 두고, 단위 변환만
`_pct`가 한다.
"""
from __future__ import annotations

import io
import math
from datetime import datetime, timedelta, timezone
from typing import Any, NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

KST = timezone(timedelta(hours=9))

Z_DEFAULT = 1.5
"""호가 폭 배수 기본값 — 프론트 `Z_DEFAULT`와 같은 1.5σ."""

HORIZON_DEFAULT = 60
"""σ_r 지평(초) 기본값 — 프론트 `QUOTE_HORIZON_SECONDS`와 같은 1분."""

FILL_ROWS = 200
"""체결 시트에 미리 깔아 두는 수식 행 수. 부족하면 마지막 행을 복사해 늘리면 된다."""

INTRADAY_CAP_WON_DEFAULT = 300_000_000
"""OMS 조건변수 C(재고한도)를 만드는 **장중 재고 상한(원)** 기본값.

O/N 상한(§14.12, 잔차 σ의 함수 = 종목별)과는 다른 축이다 — 이쪽은 *밤을 넘기지 않는* 낮 동안의
재고 한도라 잔차가 아니라 **정리 여력**(반대매매·헤지로 당일 안에 털 수 있는 규모)이 제약이고,
2026-08-27 실전 분석에서 2.5~3억 구간이 최적이었다. 종목별로 갈리지 않으므로 전 종목 동일 금액을
쓰고, 필요하면 `?intraday_cap_won=`으로 그날 값을 바꿔 받는다.
"""

LOT_SHARES = 100
"""C(재고한도)를 끊는 단위 주 수. 상한을 넘지 않게 **내림**한다."""

OMS_PCT_DIGITS = 3
"""조건변수 A·B·D의 소수 자리. OMS 변수 입력이 **5자리 제한**이라 `0.084`가 상한이다
(전략 배포 체크리스트 #5). bp를 100으로 나누면 0.001% = 0.1bp 해상도 — 호가 1틱(5원,
2만원 ETF 기준 2.5bp)보다 훨씬 곱아서 반올림으로 밴드가 흔들리지 않는다.

편향의 원점은 0 붙박이고, 자전은 함수 안 가드가 아니라 **OMS 자전방지 플래그**가 전담한다.
"""

SKEW_SIGMA_MULT = 1.0
"""OMS 시트 D(스큐 %) = 이 배수 × σ결합 (전략 v1.5, 2026-09-08).

v1.4의 정리 앵커·임계(E·F·G)는 전부 폐지 — 사용자 결정: "임계 같은 거 말고 수량에 따라
자연스럽게 가중". 호가는 항상 진입 밴드 − 편향(D×포지션/C) 하나로, 재고가 커질수록 연속으로
안쪽에 온다. 1.0인 이유(2026-09-09, 실운영 이틀째 조정): 2.0은 절반 재고에서 이미 μ+0.5σ에 정리 호가가
앉아 "한도 절반도 안 됐는데 정리가 세다"(사용자) → 1σ로 완화. 이제 한도 C에서 μ+0.5σ(터치
백테스트 청산 레벨), 절반에서 μ+1σ, 소량은 밴드 근처. 정리를 세게 하고 싶으면 D를 키운다
(마감 스텝업 14:30 ×2 / 14:45 ×3도 D 배수).
"""

MULT_K200 = 50_000      # 미니K200 승수 (§14.4)
MULT_KQ150 = 10_000     # KQ150F 승수

SHEET_BETA = "베타"
SHEET_OMS = "OMS"
SHEET_FILLS = "체결"
SHEET_HEDGE = "헤지"
SHEET_PARAMS = "PARAMS"

# PARAMS 메타 블록에서 승수 항목을 찾을 라벨 — 헤지 시트가 참조할 셀 주소를 **레이아웃에서 역산**
# 하기 위한 키다. 행 번호를 상수로 박으면 메타 줄이 하나 늘 때 헤지 수식이 조용히 엉뚱한 칸을
# 가리킨다 (숫자를 두 곳에 적지 않는다는 원칙의 연장).
_LABEL_MULT_K200 = "미니K200 승수"
_LABEL_MULT_KQ150 = "KQ150F 승수"

# ── 서식 ──────────────────────────────────────────────────────────────────
# 웹 화면 팔레트(다크)가 아니라 **엑셀 관례**(밝은 배경 + 진한 헤더)를 따른다 — 이 파일은
# 내부망 엑셀에서 그대로 열리고 인쇄도 될 수 있다.
_HEADER_FILL = PatternFill(patternType="solid", fgColor="1F2430")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_INPUT_FILL = PatternFill(patternType="solid", fgColor="FFF2CC")   # 입력/DDE 셀 (노랑)
_CALC_FILL = PatternFill(patternType="solid", fgColor="EDF2F7")    # 수식 결과 셀 (연회색)
_ACTION_FILL = PatternFill(patternType="solid", fgColor="DDF0D5")  # 집행할 계약 (연초록)
_TITLE_FONT = Font(bold=True, size=12)
_NOTE_FONT = Font(size=9, color="7F7F7F")
_LABEL_FONT = Font(bold=True, size=10)
_ACTION_FONT = Font(bold=True, size=14)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_TEXT = "@"
_FMT_BETA = "0.0000"
_FMT_BP = "0.00"
_FMT_PCT = "0.000"   # OMS 조건변수 % — 0으로 채워 3자리를 항상 보이게 (0.08 ≠ 0.084 오타 방지)
_FMT_SIGMA = "0.0"   # OMS 조건변수 E — %가 아니라 σ 배수. 1 ≠ 1.0 혼동 없게 한 자리 고정
_FMT_INT = "#,##0"
_FMT_WON = "#,##0"


def _num(v: Any) -> float | None:
    """JSON에서 온 값 → 유한 float. None/NaN/문자열은 None (셀을 비운다)."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if math.isfinite(f) else None


def _round(v: float | None, digits: int) -> float | None:
    return None if v is None else round(v, digits)


def _pct(bp: float | None) -> float | None:
    """bp → % (÷100, 소수 `OMS_PCT_DIGITS`자리) — OMS 조건변수 입력 단위."""
    return None if bp is None else round(bp / 100.0, OMS_PCT_DIGITS)


def horizon_label(seconds: int) -> str:
    """지평 라벨 — `60 → "1분"` (프론트 `horizonLabel`과 같은 표기)."""
    return f"{seconds // 60}분"


class QuoteBand(NamedTuple):
    """호가 밴드 한 벌 — 베타 시트와 OMS 시트가 **같은 계산 결과**를 나눠 쓴다.

    두 시트가 각자 μ·σ를 다시 조합하면 반올림 자리 하나만 어긋나도 "엑셀 안에서 값이 안 맞는"
    상태가 된다. 산식은 `quote_band_bp` 한 곳에만 둔다.
    """

    ask: float | None        # x매도 = OMS A (진입 매도)
    bid: float | None        # x매수 = OMS B (진입 매수)
    mu: float | None         # μ_g — 밴드 중심 (OMS 함수가 (A+B)/2로 복원하는 그 값)
    sigma_r: float | None    # σ선물 (지평 T)
    sigma_comb: float | None  # √(σ_g² + σ_r²) — 밴드 반폭의 1σ = OMS D


def quote_band_bp(calib: dict | None, z: float, horizon_seconds: int) -> QuoteBand:
    """호가 밴드 bp — 프론트 `suggestQuote`와 **같은 산식**.

    σ_comb = √(σ_g² + σ_r²), x = μ_g ± z·σ_comb. σ_r이 없으면 차단이 아니라 σ_g로 degrade
    (프론트 동일 규약 — 선물 30초봉이 없어도 호가는 서야 한다). μ_g·σ_g가 없으면 밴드 자체가 없다
    (σ_r은 알면 그대로 돌려준다 — 베타 시트의 σ선물 열은 밴드와 별개로 채워진다).
    """
    if not calib:
        return QuoteBand(None, None, None, None, None)
    mu = _num(calib.get("g_mean_bp"))
    sigma_g = _num(calib.get("g_sigma_level_bp"))
    sigma_r = _num((calib.get("s_diff_sigma_bp") or {}).get(str(horizon_seconds)))
    if mu is None or sigma_g is None:
        return QuoteBand(None, None, None, sigma_r, None)
    sigma_comb = math.hypot(sigma_g, sigma_r or 0.0)
    half = z * sigma_comb
    return QuoteBand(
        _round(mu + half, 2), _round(mu - half, 2), mu, sigma_r, _round(sigma_comb, 2)
    )


def inventory_cap_shares(prev_close: float | None, cap_won: float) -> int | None:
    """OMS c(재고한도, 주) = 장중 재고 상한 ÷ 전일종가 → **100주 내림**.

    내림인 이유는 c가 *정리 모드로 넘어가는 문턱*이라, 반올림으로 상한을 넘어서면 안 되기
    때문이다. 전일종가가 없거나 0이면 문턱을 만들 수 없다 (None → 빈칸 + 비고).
    """
    if prev_close is None or prev_close <= 0 or cap_won <= 0:
        return None
    return int(math.floor(cap_won / prev_close / LOT_SHARES)) * LOT_SHARES


# ---------------------------------------------------------------------------
# 시트 빌더
# ---------------------------------------------------------------------------


def _widths(ws: Worksheet, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws: Worksheet, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER


def overnight_label(item: dict) -> str:
    """O/N 금지 열 값 — `금지` / `표본부족` / 빈칸(허용). §14.12.

    "금지"(σ를 재 보니 크더라)와 "표본부족"(아직 못 쟀다)은 다른 상태다. 허용은 빈칸으로 둬
    금지 행만 눈에 띄게 한다 (엑셀 필터로도 바로 걸린다).
    """
    if item.get("overnight_banned"):
        return "금지"
    return "" if _num(item.get("overnight_cap_won")) is not None else "표본부족"


def _build_beta(ws: Worksheet, items: list[dict], z: float, horizon: int) -> int:
    """베타 시트 — /master 캐시 스냅샷. 반환은 기록한 데이터 행 수.

    잔차σ 바로 옆에 **O/N 상한·금지**를 붙인다 (§14.12) — 상한은 잔차σ만의 함수라, 셋이 나란히
    있어야 "이 σ라서 이 금액"이 파일에서도 읽힌다. 조회 키 열(A~D)은 건드리지 않으므로 체결
    시트 VLOOKUP은 그대로다.
    """
    labels = [
        "코드", "종목명", "β_K200", "β_KQ150", "R²", "잔차σ(bp)",
        "O/N 상한(원)", "O/N 금지",
        "μ괴리(bp)", "σ괴리(bp)", f"σ선물(bp·{horizon_label(horizon)})",
        "x매도(bp)", "x매수(bp)", "전일종가", "CU",
    ]
    _header_row(ws, 1, labels)
    _widths(ws, [10, 26, 9, 9, 8, 10, 13, 9, 11, 11, 13, 11, 11, 11, 10])

    for i, it in enumerate(items):
        row = i + 2
        calib = it.get("calib")
        band = quote_band_bp(calib, z, horizon)
        values = [
            str(it.get("etf_code") or ""),
            it.get("name") or "",
            _num(it.get("beta_k200")),
            _num(it.get("beta_kq150")),
            _num(it.get("r2")),
            _num(it.get("resid_vol_bp")),
            _num(it.get("overnight_cap_won")),
            overnight_label(it),
            _num((calib or {}).get("g_mean_bp")),
            _num((calib or {}).get("g_sigma_level_bp")),
            band.sigma_r,
            band.ask,
            band.bid,
            _num(it.get("prev_close")),
            _num(it.get("creation_unit")),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = _BORDER
        # 코드는 **텍스트 서식** — 0117V0·0052D0 같은 영숫자/선행0 코드가 숫자로 굳으면
        # 체결 시트 VLOOKUP 키가 어긋난다.
        ws.cell(row=row, column=1).number_format = _FMT_TEXT
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = _FMT_BETA
        ws.cell(row=row, column=6).number_format = _FMT_BP
        ws.cell(row=row, column=7).number_format = _FMT_WON
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
        for col in (9, 10, 11, 12, 13):
            ws.cell(row=row, column=col).number_format = _FMT_BP
        ws.cell(row=row, column=14).number_format = _FMT_WON
        ws.cell(row=row, column=15).number_format = _FMT_INT

    ws.freeze_panes = "C2"   # 헤더 + 코드·종목명 고정 (15열이라 가로 스크롤이 잦다)
    return len(items)


def oms_note(
    item: dict,
    band_pct: tuple[float | None, float | None],
    cap_shares: int | None,
) -> str:
    """OMS 행의 비고 — 그 행을 **그대로 반입하면 안 되는 이유**만 적는다 (정상 행은 빈칸).

    전략 v1.5의 제약은 `A > B` 하나다 — 간격 양수라야 매도−매수 = (A−B) > 0 으로 자기 크로스가
    산술적으로 불가하다. 산식상 A−B = 2z·σ_comb 라 정상 경로에선 안 깨지지만, σ가 0에 가까워
    %로 반올림하며 붙는 경계가 있다 — 그 행만 반입하지 않으면 된다.

    O/N 금지(§14.12)는 값을 못 쓰게 만드는 사유가 아니라 *마감 전에 반드시 턴다*는 별도 제약이라
    값과 함께 남긴다. 종가 지연은 C의 분모가 낡았다는 뜻이므로 값을 지우지 않고 표시만 한다.
    """
    notes: list[str] = []
    a_pct, b_pct = band_pct
    if a_pct is None or b_pct is None:
        notes.append("통계 없음")
    elif a_pct <= b_pct:
        notes.append("⚠ A>B 위반 반입 금지")
    if cap_shares is None:
        notes.append("전일종가 없음")
    elif item.get("prev_close_stale"):
        notes.append("전일종가 지연")
    if item.get("overnight_banned"):
        notes.append("당일 정리 필수")
    return " · ".join(notes)


def _build_oms(ws: Worksheet, items: list[dict], z: float, horizon: int, cap_won: float) -> None:
    """OMS 시트 — 함수 전략 v1.5의 조건변수 **A~D**를 종목별로 옮겨 적는 표 (§14.11).

    전략 정본은 `docs/lp-oms-strategy-v1.html` (v1.5, 2026-09-08 확정). OMS 함수 3개가 이 4개
    변수만 읽어 호가를 만든다 — 분기 없는 연속 스큐 단일 구조:

        편향   = RealNav × D × 포지션 / C / 100
        매도   = RealNav × (100 + A) / 100 − 편향
        매수   = RealNav × (100 + B) / 100 − 편향

    A·B는 베타 시트의 x매도/x매수와 같은 함수(`quote_band_bp`)의 bp ÷ 100, D는 그 밴드의
    σ결합 × `SKEW_SIGMA_MULT`(1.0) — 한도 C에서 호가가 μ±0.5σ(백테스트 청산 레벨)까지 밀린다.
    이력: v1.2 정리레벨 E·F% → v1.3 (A+B)/2 복원+강도 E → v1.4 F=μ 직접+임계 G → v1.5 전부
    폐지·연속 스큐 일원화(사용자 결정). **A~D 4칸만** 초록, 참고 열(μ·전일종가·상한)과 분리.
    """
    labels = [
        "코드", "종목명",
        "A 매도(%)", "B 매수(%)", "C 재고한도(주)", f"D 스큐(%·{horizon_label(horizon)}·1σ)",
        "(참고) μ괴리(%)", "전일종가", "장중상한(원)", "비고",
    ]
    _header_row(ws, 1, labels)
    _widths(ws, [10, 26, 11, 11, 13, 14, 13, 11, 15, 22])

    for i, it in enumerate(items):
        row = i + 2
        band = quote_band_bp(it.get("calib"), z, horizon)
        cap_shares = inventory_cap_shares(_num(it.get("prev_close")), cap_won)
        a_pct, b_pct = _pct(band.ask), _pct(band.bid)
        values = [
            str(it.get("etf_code") or ""),
            it.get("name") or "",
            a_pct,
            b_pct,
            cap_shares,
            _pct(band.sigma_comb and band.sigma_comb * SKEW_SIGMA_MULT),
            _pct(band.mu),
            _num(it.get("prev_close")),
            cap_won,
            oms_note(it, (a_pct, b_pct), cap_shares),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = _BORDER
        ws.cell(row=row, column=1).number_format = _FMT_TEXT
        for col in (3, 4, 6, 7):
            ws.cell(row=row, column=col).number_format = _FMT_PCT
        ws.cell(row=row, column=5).number_format = _FMT_INT
        for col in (8, 9):
            ws.cell(row=row, column=col).number_format = _FMT_WON
        # A~D = **OMS에 실제로 넣는 4칸**. 참고 열과 색으로 갈라 놓아야 옮겨 적을 때 헷갈리지 않는다.
        for col in range(3, 7):
            ws.cell(row=row, column=col).fill = _ACTION_FILL

    ws.freeze_panes = "C2"


def _build_fills(ws: Worksheet, beta_rows: int) -> None:
    """체결 시트 — 입력/DDE 200행 + VLOOKUP 수식.

    빈 행에서 `#N/A`·`0`이 깔리면 시트를 못 읽는다. 이름·β는 IFERROR로, 금액 3열은 입력이
    비면 `""`로 비운다 (헤지 시트의 SUM은 텍스트를 무시하므로 합계에 영향 없음).
    코드 조회 키는 `$A2&""` — 베타 시트의 코드는 텍스트인데 DDE·수기 입력이 `396500`을 **숫자**로
    넣으면 VLOOKUP이 조용히 `#N/A`가 된다. 빈 문자열 결합은 숫자·텍스트·빈칸을 전부 텍스트로
    통일하는 가장 이식성 높은 관용구다 (`TEXT(...,"@")`는 엔진에 따라 숫자에서 어긋난다 — 실측).
    조회 범위는 열 전체가 아니라 **베타 데이터 행으로 한정**한다 (200행 × 3 VLOOKUP이 100만 행을
    훑지 않게 + 헤더행 오매칭 원천 차단). 파일은 매번 서버가 다시 만들므로 범위 갱신은 자동이다.
    """
    labels = [
        "코드", "종목명", "수량", "현재가", "평가액",
        "β_K200", "β_KQ150", "K200 노출", "KQ150 노출",
    ]
    _header_row(ws, 1, labels)
    _widths(ws, [11, 26, 12, 12, 16, 9, 9, 16, 16])

    # A=코드 / B=종목명 / C=β_K200 / D=β_KQ150 (베타 시트 열 순서)
    lookup = f"{SHEET_BETA}!$A$2:$D${beta_rows + 1}"
    for row in range(2, FILL_ROWS + 2):
        key = f'$A{row}&""'
        ws.cell(row=row, column=1).number_format = _FMT_TEXT
        ws.cell(row=row, column=2, value=f'=IFERROR(VLOOKUP({key},{lookup},2,0),"")')
        ws.cell(row=row, column=3).number_format = _FMT_INT
        ws.cell(row=row, column=4).number_format = _FMT_WON
        e = ws.cell(row=row, column=5, value=f'=IF(OR($C{row}="",$D{row}=""),"",$C{row}*$D{row})')
        e.number_format = _FMT_WON
        f = ws.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP({key},{lookup},3,0),0)')
        g = ws.cell(row=row, column=7, value=f'=IFERROR(VLOOKUP({key},{lookup},4,0),0)')
        f.number_format = g.number_format = _FMT_BETA
        h = ws.cell(row=row, column=8, value=f'=IF($E{row}="","",$E{row}*$F{row})')
        i = ws.cell(row=row, column=9, value=f'=IF($E{row}="","",$E{row}*$G{row})')
        h.number_format = i.number_format = _FMT_WON
        for col in (1, 3, 4):
            ws.cell(row=row, column=col).fill = _INPUT_FILL
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = _BORDER

    ws.freeze_panes = "A2"


def _build_hedge(ws: Worksheet, mult_cells: tuple[str, str]) -> None:
    """헤지 시트 — 노출 합계 → 목표 계약수 → **집행할 계약**.

    선물가가 비어 있으면 `ROUND(-노출/(가격×승수),0)`이 `#DIV/0!`가 되므로 `N()` 가드로 막는다
    (`N()`은 공백·텍스트를 0으로 준다 — 미입력/문자 모두 한 번에 걸린다).
    `mult_cells`는 PARAMS 시트가 알려준 승수 셀 주소 — 승수 숫자는 이 워크북에 한 벌만 존재한다.
    """
    mult_k200, mult_kq150 = mult_cells
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="지수선물 헤지 계산 — 노출 → 목표 계약수 → 집행")
    t.font = _TITLE_FONT
    ws.cell(row=2, column=1, value="노란 셀만 입력(또는 DDE 연결). 나머지는 수식 — 부호는 음수=매도(숏), 양수=매수.").font = _NOTE_FONT
    _widths(ws, [30, 20, 20])

    _header_row(ws, 4, ["항목", "미니K200", "KQ150F"])

    def line(row: int, label: str, bk: str | None, bq: str | None, fmt: str,
             *, input_cell: bool = False, calc: bool = False) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _LABEL_FONT
        lc.border = _BORDER
        for col, formula in ((2, bk), (3, bq)):
            c = ws.cell(row=row, column=col, value=formula)
            c.number_format = fmt
            c.border = _BORDER
            c.alignment = Alignment(horizontal="right")
            if input_cell:
                c.fill = _INPUT_FILL
            elif calc:
                c.fill = _CALC_FILL

    line(5, "ETF 노출 (₩)", f"=SUM({SHEET_FILLS}!H:H)", f"=SUM({SHEET_FILLS}!I:I)", _FMT_WON, calc=True)
    line(6, "선물 현재가  ← 입력/DDE", None, None, _FMT_WON, input_cell=True)
    line(7, "승수 (PARAMS)", f"={mult_k200}", f"={mult_kq150}", _FMT_INT)
    line(8, "1계약 명목 (₩)", '=IF(N(B6)=0,"",B6*B7)', '=IF(N(C6)=0,"",C6*C7)', _FMT_WON, calc=True)
    line(9, "목표 계약수", '=IF(N(B6)=0,"",ROUND(-B5/(B6*B7),0))', '=IF(N(C6)=0,"",ROUND(-C5/(C6*C7),0))', _FMT_INT, calc=True)
    line(10, "보유 계약수  ← 입력/DDE", None, None, _FMT_INT, input_cell=True)
    line(11, "집행할 계약 (목표−보유)", '=IF(B9="","",B9-N(B10))', '=IF(C9="","",C9-N(C10))', _FMT_INT)
    line(12, "헤지 후 잔여 노출 (₩)", '=IF(N(B6)=0,"",B5+N(B10)*B6*B7)', '=IF(N(C6)=0,"",C5+N(C10)*C6*C7)', _FMT_WON, calc=True)

    # 집행할 계약 = 이 시트의 결론. 크게.
    for col in (2, 3):
        c = ws.cell(row=11, column=col)
        c.font = _ACTION_FONT
        c.fill = _ACTION_FILL
    ws.row_dimensions[11].height = 24

    ws.cell(row=14, column=1, value="목표 계약수 = ROUND( −ETF 노출 ÷ (선물 현재가 × 승수), 0 )").font = _NOTE_FONT
    ws.cell(row=15, column=1, value="ETF를 매수(노출 +)했으면 목표는 음수 = 선물 매도로 헤지한다.").font = _NOTE_FONT
    ws.cell(row=16, column=1, value="선물 현재가가 비어 있으면 목표·집행 칸은 빈칸으로 남는다 (0으로 나누지 않는다).").font = _NOTE_FONT


def _build_params(
    ws: Worksheet, master: dict, z: float, horizon: int, rows: int, cap_won: float
) -> tuple[str, str]:
    """PARAMS 시트 — 스냅샷 메타 + 승수(헤지 시트 참조원) + 사용법.

    반환은 (미니K200 승수 셀, KQ150F 승수 셀)의 절대 주소 — 헤지 시트가 이 주소를 참조한다.
    행 번호를 상수로 박지 않고 여기서 역산해 넘겨야, 메타 줄이 늘거나 순서가 바뀌어도 헤지
    수식이 엉뚱한 칸을 가리키지 않는다.
    """
    cp = master.get("calib_params") or {}
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="LENS LP 데스크 — 파라미터 스냅샷").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="매크로 없음 · 순수 수식. 외부망 LENS에서 생성 → 내부망 엑셀에서 DDE와 결합.").font = _NOTE_FONT
    _widths(ws, [26, 34])

    meta: list[tuple[str, Any, str | None]] = [
        ("생성시각 (KST)", datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"), None),
        ("통계일 (stats_date)", master.get("stats_date") or "", None),
        ("캘리브 as_of", cp.get("as_of") or "", None),
        ("캘리브 창 (거래일)", _num(cp.get("calib_days")), _FMT_INT),
        ("호가 z (σ 배수)", z, "0.00"),
        ("호가 지평 T (초)", horizon, _FMT_INT),
        ("장중 재고 상한 (원)", cap_won, _FMT_WON),
        (_LABEL_MULT_K200, MULT_K200, _FMT_INT),
        (_LABEL_MULT_KQ150, MULT_KQ150, _FMT_INT),
        ("유니버스 종목수", rows, _FMT_INT),
        ("g 표본 구간", cp.get("g_window") or "-", None),
        ("s 표본 구간", cp.get("s_window") or "-", None),
    ]
    mult_cells: dict[str, str] = {}
    for i, (label, value, fmt) in enumerate(meta):
        row = i + 3
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _LABEL_FONT
        lc.border = _BORDER
        c = ws.cell(row=row, column=2, value=value)
        c.border = _BORDER
        if fmt:
            c.number_format = fmt
        if label in (_LABEL_MULT_K200, _LABEL_MULT_KQ150):
            mult_cells[label] = f"{SHEET_PARAMS}!$B${row}"

    # 오버나이트 상한 룰(§14.12)은 **숫자를 여기 적지 않는다** — `/master`의 params가 단일
    # 진실원이고, 그게 없으면 룰 설명 자체를 빼서 반쪽짜리 문장이 나가지 않게 한다.
    p = master.get("params") or {}
    on = [_num(p.get(k)) for k in ("on_tail_loss_won", "on_tail_z", "on_max_resid_vol_bp", "on_min_obs")]
    on_lines: list[str] = []
    if all(v is not None for v in on):
        tail, tz, max_sigma, min_obs = on
        on_lines = [
            f"· O/N 상한 = {tail:,.0f} ÷ ({tz:g} × 잔차σ(bp) ÷ 10,000) — 1박 5% 꼬리손실이 {tail:,.0f}원을 넘지 않는 평가액(백만 내림).",
            f"  잔차σ > {max_sigma:g}bp 또는 회귀 표본 {min_obs:,.0f}일 미만이면 'O/N 금지'(상한 0) — 넘기지 말고 당일 정리한다.",
        ]

    usage = [
        "",
        "■ 사용법",
        "1. [OMS] 종목별 조건변수 A~D(초록 4칸)를 장 시작 전 OMS 조건변수에 그대로 입력한다 (A·B·D는 % 단위·부호 그대로).",
        "2. [체결] A열 코드 · C열 수량 · D열 현재가 — 이 3개만 입력하거나 DDE에 연결한다.",
        "   나머지 열(B·E~I)은 수식이다. 200행까지 준비돼 있고, 부족하면 마지막 행을 복사해 늘린다.",
        "3. [헤지] 선물 현재가(B6·C6)와 보유 계약수(B10·C10)를 입력하거나 DDE에 연결한다.",
        "4. [헤지] '집행할 계약'이 곧 주문 수량이다 — 음수 = 매도(숏), 양수 = 매수.",
        "",
        "■ [OMS] 조건변수 A~D — 함수 전략 v1.5 (docs/lp-oms-strategy-v1.html)",
        "  OMS 함수 3개(70 매도호가 · 71 매수호가 · 72 재고편향)가 이 4개만 읽는다. 표의 값을 그대로 입력.",
        "  분기·임계·정리 앵커 없음 — 호가는 항상 진입 밴드 − 편향 하나다:",
        "    편향 = RealNav × D × 포지션 / C / 100",
        "    매도 = RealNav × (100 + A) / 100 − 편향  ·  매수 = RealNav × (100 + B) / 100 − 편향",
        "  (두 전략은 독립 할당 — 같은 편향을 빼므로 어느 재고에서도 매도 − 매수 = A−B 로 간격 고정.)",
        f"· A 매도(%)·B 매수(%) = 진입 밴드 = μ괴리 ± {z:g}σ결합을 %로 옮긴 값 — [베타] 시트의 x매도/x매수(bp) ÷ 100이다"
        " (z는 위 '호가 z'로 바꿔 받는다). 플랫일 때 양쪽이 여기 선다.",
        "  부호는 그대로 쓴다 (매도 +5bp → 0.05, 매수 −30bp → −0.30). 소수 3자리 — OMS 변수 5자리 제한 안이다.",
        "  장중 수동 조정 자유 — A·B는 진입 위치만 정하고, 정리는 편향이 재고에 비례해 자동으로 이끈다.",
        f"· C 재고한도(주) = 장중 재고 상한 ÷ 전일종가, 100주 내림 (상한은 위 '장중 재고 상한' {cap_won:,.0f}원 — 전 종목 동일).",
        "  험한 날(종목 간 등락 격차 평소 1.5배+)은 C를 절반으로: LENS에서 ?intraday_cap_won= 를 반값으로 다시 받는다.",
        f"· D 스큐(%) = σ결합 × {SKEW_SIGMA_MULT:g} — 한도 C에서 밀리는 총 폭. 재고 비율만큼만 내려온다:",
        "  매도호가가 μ+0.5σ(한도 = 터치 백테스트 청산 레벨) · μ+1σ(절반) · 밴드 근처(<25%)에 자연 연속으로 선다.",
        "  정리가 느리면 D를 키운다 — 임계·앵커 없이 이 숫자 하나가 정리 속도다.",
        "· 제약은 하나 — A > B (간격 양수 → 자기 크로스 산술적 불가). 위반 행은 비고에 '⚠ A>B 위반 반입 금지'.",
        "⚠️ 전 주문에 '자전방지 플래그'를 활성화한다 — 시차 자전(앵커 분기 전환 시 옛 반대편 주문과 만나는 경로)은 이 플래그가 전담한다.",
        "· 마감 전 정리는 D 스텝업 (14:30 ×2 · 14:45 ×3), 14:55 잔량은 CD 시분할. 전 종목 일괄 변경이면 되고 함수 수정은 없다.",
        "· 구 조건변수 a~e(재고 c에서 정리 호가 d·e로 스위칭)는 폐지됐다 — 정리는 연속 편향(상시) + D 스텝업(마감 전)이 대체한다.",
        "· [OMS] 비고 '당일 정리 필수' = O/N 금지 종목(§14.12). A~D는 그대로 쓰되 마감 전 반드시 턴다.",
        "  '통계 없음'은 캘리브가 없어 A·B·D가 빈칸인 행 — 그 종목은 수기 판단이다(C는 전일종가만 있으면 나온다).",
        "",
        "■ 값의 의미",
        "· β_K200/β_KQ150 = ETF 일간수익률의 2-팩터 회귀 계수 (K200·KQ150 지수, 창 120영업일).",
        "  노출 = 평가액 × β 이므로, 체결한 ETF 금액에 β를 곱한 것이 지수 환산 델타다.",
        f"· x매도/x매수 = iNAV 대비 호가 밴드(bp). x = μ괴리 ± {z:g}σ, σ = √(σ괴리² + σ선물²).",
        "  매도 호가 = iNAV × (1 + x매도/10000), 매수 호가 = iNAV × (1 + x매수/10000).",
        "  [OMS] 시트의 A·B·D는 같은 값을 ÷100 해 % 로 옮긴 것이다 — 두 시트가 어긋나면 한쪽이 낡은 파일이다.",
        *on_lines,
        "· 전일종가·CU는 생성 시점 스냅샷이다.",
        "· [체결] 코드가 [베타]에 없으면 종목명이 비고 β가 0으로 잡힌다(노출 0). 종목명 칸이 검산 포인트다.",
        "",
        "■ 주의",
        "· 이 파일의 파라미터는 매일 바뀐다. 장 시작 전 LENS에서 새로 받아 쓴다.",
        "· [베타] 코드 열은 텍스트 서식이다 (0117V0·0052D0 같은 영숫자 코드). 숫자로 바꾸지 말 것.",
        "· 시트 이름을 바꾸면 수식이 깨진다 (베타/체결/헤지/PARAMS).",
    ]
    for i, line in enumerate(usage):
        c = ws.cell(row=len(meta) + 4 + i, column=1, value=line)
        c.font = _LABEL_FONT if line.startswith(("■", "⚠️")) else _NOTE_FONT

    return mult_cells[_LABEL_MULT_K200], mult_cells[_LABEL_MULT_KQ150]


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------


def build_workbook(
    master: dict,
    *,
    z: float = Z_DEFAULT,
    horizon: int = HORIZON_DEFAULT,
    intraday_cap_won: float = INTRADAY_CAP_WON_DEFAULT,
) -> bytes:
    """`/master` 페이로드 → xlsx 바이트. 순수 함수 (PG·전역 상태 접근 없음)."""
    items = list(master.get("items") or [])
    wb = Workbook()
    # 시트는 먼저 다 만들어 **탭 순서**를 고정하고(생성 순서 = 탭 순서), 채우기는 의존 순서대로:
    # PARAMS가 승수 셀 주소를 확정해야 헤지 수식을 쓸 수 있다.
    beta = wb.active
    beta.title = SHEET_BETA
    oms = wb.create_sheet(SHEET_OMS)
    fills = wb.create_sheet(SHEET_FILLS)
    hedge = wb.create_sheet(SHEET_HEDGE)
    params = wb.create_sheet(SHEET_PARAMS)

    rows = _build_beta(beta, items, z, horizon)
    _build_oms(oms, items, z, horizon, intraday_cap_won)
    _build_fills(fills, rows)
    _build_hedge(hedge, _build_params(params, master, z, horizon, rows, intraday_cap_won))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(now: datetime | None = None) -> str:
    return f"lp_desk_params_{(now or datetime.now(KST)):%Y%m%d}.xlsx"
